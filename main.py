import advertools as adv
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from polyfuzz import PolyFuzz
from polyfuzz.models import RapidFuzz

matcher = RapidFuzz(n_jobs=1, score_cutoff=0.85)
model = PolyFuzz(matcher)

st.markdown("""
<style>
.big-font {
    font-size:50px !important;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<p class="big-font">SEO URL Redirect Mapper</p>
<b>Directions: </b>
<ul>
<li>Upload Legacy Crawl or URLs (xlsx)</li>
<li>Upload New Crawl or URLs (xlsx)</li>
</ul>
<b>Requirements: </b>
<ul>
<li>Column 1 to be named "Address" and contain full URLs, including http(s)://</li>
<li>The following column headings need to exist in both files, even if column cells are blank:
 <ul>
 <li>"Title-1" "H1-1" "H2-1"</li>
 </ul>

</ul>
""", unsafe_allow_html=True)

legacy_file = st.file_uploader('Upload Crawl of LEGACY URLs', type='xlsx', key='legacy')
# legacy_file = 'tl_products_internal_all.xlsx'

input_files = []
crawl_columns = ['Address', 'Title 1', 'H1-1', 'H2-1']


def analyze_crawls(crawls):
    st.write('Reading site crawls. . .')
    for crawl in crawls:
        wb = load_workbook(filename=crawl)
        sheet_name = wb.sheetnames
        input_files.append([crawl, sheet_name])

    # Read Files and Gather Column Names
    legacy_crawl = pd.read_excel(input_files[0][0], sheet_name=input_files[0][1][0])
    legacy_crawl = legacy_crawl[crawl_columns]
    new_crawl = pd.read_excel(input_files[1][0], sheet_name=input_files[1][1][0])
    new_crawl = new_crawl[crawl_columns]
    legacy_col_1 = list(legacy_crawl.columns)[0]
    new_col_1 = list(new_crawl.columns)[0]
    legacy_urls = legacy_crawl[legacy_col_1].tolist()
    new_urls = new_crawl[new_col_1].tolist()
    st.write(legacy_crawl[legacy_col_1].head())
    st.write(new_crawl[new_col_1].head())

    # st.write(list(legacy_crawl.columns))
    # st.write(list(new_crawl.columns))
    url_parse(legacy_urls, legacy_crawl, new_urls, new_crawl)


def url_match(legacy_paths, new_paths, legacy_url_parse, new_url_parse):
    st.write('Analyzing URL Paths. . .')
    model.match(legacy_paths, new_paths)

    pfuzz_df = model.get_matches()
    pfuzz_df["Similarity"] = pfuzz_df["Similarity"].round(3)
    pfuzz_df = pfuzz_df.sort_values('Similarity', ascending=False)
    pfuzz_df = pfuzz_df[pfuzz_df['Similarity'] >= .900]
    print(pfuzz_df.head())

    join_df = pd.merge(pfuzz_df, legacy_url_parse, left_on='From', right_on='path')
    join_df_2 = pd.merge(join_df, new_url_parse, left_on='To', right_on='path')
    join_df_2.rename(
        columns={'url_x': 'Legacy URL', 'url_y': 'New URL', 'path_x': 'Legacy URL Path', 'path_y': 'New URL Path'},
        inplace=True)
    url_df = join_df_2[['From', 'To', 'Similarity', 'Legacy URL Path', 'New URL Path', 'Legacy URL', 'New URL']]
    url_df = url_df.drop_duplicates()
    url_df.head()
    st.dataframe(url_df)
    return url_df


def slug_match(legacy_slugs, new_slugs, legacy_url_parse, new_url_parse):
    st.write('Analyzing URL Slugs. . .')
    model.match(legacy_slugs, new_slugs)

    pfuzz_df = model.get_matches()
    pfuzz_df["Similarity"] = pfuzz_df["Similarity"].round(3)
    pfuzz_df = pfuzz_df.sort_values('Similarity', ascending=False)
    pfuzz_df = pfuzz_df[pfuzz_df['Similarity'] >= .800]
    print(pfuzz_df.head())

    join_df = pd.merge(pfuzz_df, legacy_url_parse, left_on='From', right_on='last_dir')
    join_df_2 = pd.merge(join_df, new_url_parse, left_on='To', right_on='last_dir')
    join_df_2.rename(
        columns={'url_x': 'Legacy URL', 'url_y': 'New URL', 'path_x': 'Legacy URL Path', 'path_y': 'New URL Path'},
        inplace=True)
    slug_df = join_df_2[['From', 'To', 'Similarity', 'Legacy URL Path', 'New URL Path', 'Legacy URL', 'New URL']]
    slug_df = slug_df.drop_duplicates()
    slug_df.head()
    st.dataframe(slug_df)
    return slug_df

def title_match(legacy_titles, new_titles, legacy_crawl, new_crawl):
    st.write('Analyzing Title tags. . .')
    model.match(legacy_titles, new_titles)

    pfuzz_df = model.get_matches()
    pfuzz_df["Similarity"] = pfuzz_df["Similarity"].round(3)
    pfuzz_df = pfuzz_df.sort_values('Similarity', ascending=False)
    pfuzz_df = pfuzz_df[pfuzz_df['Similarity'] >= .900]
    print(pfuzz_df.head())

    join_df = pd.merge(pfuzz_df, legacy_crawl, left_on='From', right_on='Title 1')
    join_df_2 = pd.merge(join_df, new_crawl, left_on='To', right_on='Title 1').drop_duplicates()
    join_df_2.rename(columns={'Address_x': 'Legacy URL', 'Address_y': 'New URL'}, inplace=True)
    title_df = join_df_2[['From', 'To', 'Similarity', 'Legacy URL', 'New URL']]
    title_df = title_df.drop_duplicates()
    print(title_df.head())
    st.dataframe(title_df)
    return title_df


def h1_match(legacy_h1, new_h1, legacy_crawl, new_crawl):
    st.write('Analyzing h1 tags. . .')
    model.match(legacy_h1, new_h1)

    pfuzz_df = model.get_matches()
    pfuzz_df["Similarity"] = pfuzz_df["Similarity"].round(3)
    pfuzz_df = pfuzz_df.sort_values('Similarity', ascending=False)
    pfuzz_df = pfuzz_df[pfuzz_df['Similarity'] >= .900]
    print(pfuzz_df.head())

    join_df = pd.merge(pfuzz_df, legacy_crawl, left_on='From', right_on='H1-1')
    join_df_2 = pd.merge(join_df, new_crawl, left_on='To', right_on='H1-1')
    join_df_2.rename(columns={'Address_x': 'Legacy URL', 'Address_y': 'New URL'}, inplace=True)
    h1_df = join_df_2[['From', 'To', 'Similarity', 'Legacy URL', 'New URL']]
    h1_df = h1_df.drop_duplicates()
    st.dataframe(h1_df)
    return h1_df


def h2_match(legacy_h2, new_h2, legacy_crawl, new_crawl):
    st.write('Analyzing h2 tags. . .')
    model.match(legacy_h2, new_h2)

    pfuzz_df = model.get_matches()
    pfuzz_df["Similarity"] = pfuzz_df["Similarity"].round(3)
    pfuzz_df = pfuzz_df.sort_values('Similarity', ascending=False)
    pfuzz_df = pfuzz_df[pfuzz_df['Similarity'] >= .900]
    print(pfuzz_df.head())

    join_df = pd.merge(pfuzz_df, legacy_crawl, left_on='From', right_on='H2-1')
    join_df_2 = pd.merge(join_df, new_crawl, left_on='To', right_on='H2-1').drop_duplicates()
    join_df_2.rename(columns={'Address_x': 'Legacy URL', 'Address_y': 'New URL'}, inplace=True)
    h2_df = join_df_2[['From', 'To', 'Similarity', 'Legacy URL', 'New URL']]
    h2_df = h2_df.drop_duplicates()
    print(h2_df.head())
    st.dataframe(h2_df)
    return h2_df


def url_parse(legacy_urls, legacy_crawl, new_urls, new_crawl):
    st.write('Deconstructing URLs. . .')
    url_parse_cols = ['url', 'path', 'last_dir']
    legacy_url_parse = adv.url_to_df(legacy_urls)
    legacy_url_parse = legacy_url_parse[url_parse_cols]
    new_url_parse = adv.url_to_df(new_urls)
    new_url_parse = new_url_parse[url_parse_cols]

    legacy_paths = legacy_url_parse['path']
    new_paths = new_url_parse['path']
    legacy_slug = legacy_url_parse['last_dir']
    new_slug = new_url_parse['last_dir']
    legacy_titles = legacy_crawl['Title 1']
    new_titles = new_crawl['Title 1']
    legacy_h1 = legacy_crawl['H1-1']
    new_h1 = new_crawl['H1-1']
    legacy_h2 = legacy_crawl['H2-1']
    new_h2 = new_crawl['H2-1']
    match_dfs = [
        url_match(legacy_paths, new_paths, legacy_url_parse, new_url_parse),
        slug_match(legacy_slug, new_slug, legacy_url_parse, new_url_parse),
        title_match(legacy_titles, new_titles, legacy_crawl, new_crawl),
        h1_match(legacy_h1, new_h1, legacy_crawl, new_crawl),
        h2_match(legacy_h2, new_h2, legacy_crawl, new_crawl)
    ]

    export_dfs(match_dfs)


def export_dfs(match_dfs):
    sheet_names = ['URL Match', 'Slug Match', 'Title Match', 'H1 Match', 'H2 Match']
    with pd.ExcelWriter('mapped_urls.xlsx') as writer:
        for df in enumerate(match_dfs):
            print(df[1])
            df[1].to_excel(writer, sheet_name=sheet_names[df[0]], index=False)

    my_file = pd.read_excel('mapped_urls.xlsx')

    with open("mapped_urls.xlsx", "rb") as file:
        st.download_button(label='Download Match Analysis',
                           data=file,
                           file_name='mapped_urls.xlsx',
                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
    if legacy_file is not None:
        new_file = st.file_uploader('Upload Crawl of NEW URLs', type='xlsx', key='new')
        # new_file = 'circa_lighting_us_internal_html.xlsx'
        if new_file is not None:
            crawl_files = [legacy_file, new_file]
            analyze_crawls(crawl_files)

st.write('Author: [Tyler Gargula](https://tylergargula.dev) | Technical SEO & Software Developer')
